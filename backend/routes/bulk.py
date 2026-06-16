from flask import Blueprint, request, jsonify, send_file
from db import supabase
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.datavalidation import DataValidation
import io

bulk_bp = Blueprint("bulk", __name__)


@bulk_bp.get("/template")
def download_template():
    # Fetch cohort names from DB
    cohorts = supabase.table("cohorts").select("name").execute().data
    cohort_names = [c["name"] for c in cohorts] if cohorts else ["Future Cohort"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Hidden sheet for dropdown data
    ref_ws = wb.create_sheet("_ref", 1)
    ref_ws.sheet_state = "hidden"

    # Write dropdown options to hidden sheet
    business_types = ["Retail", "Consulting", "Manufacturing", "Education", "Healthcare", "Real Estate", "IT/Tech", "Food & Beverage", "Other"]
    team_sizes = ["1-10", "11-50", "51-200", "200+"]
    sources = ["Instagram", "Facebook", "Referral", "Cold Call", "WhatsApp", "Website", "Event", "Other"]

    for i, v in enumerate(business_types, 1): ref_ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(team_sizes, 1): ref_ws.cell(row=i, column=2, value=v)
    for i, v in enumerate(sources, 1): ref_ws.cell(row=i, column=3, value=v)
    for i, v in enumerate(cohort_names, 1): ref_ws.cell(row=i, column=4, value=v)

    # Headers
    headers = ["name *", "phone *", "whatsapp", "business_type", "city", "team_size", "source", "cohort"]
    header_fill = PatternFill("solid", fgColor="00D4A0")
    req_fill = PatternFill("solid", fgColor="FF6B35")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="000000", size=11)
        cell.fill = req_fill if "*" in h else header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [25, 18, 18, 22, 18, 12, 18, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Sample row
    sample = ["Rajesh Sharma", "9876543210", "9876543210", "Retail", "Mumbai", "1-10", "Instagram", cohort_names[0]]
    for col, val in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=val)

    # Data validations (rows 2-1001)
    def make_dv(formula):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
        dv.error = "Please select from the list"
        dv.errorTitle = "Invalid value"
        return dv

    bt_last = len(business_types)
    ts_last = len(team_sizes)
    src_last = len(sources)
    coh_last = len(cohort_names)

    dv_bt = make_dv(f"_ref!$A$1:$A${bt_last}")
    dv_ts = make_dv(f"_ref!$B$1:$B${ts_last}")
    dv_src = make_dv(f"_ref!$C$1:$C${src_last}")
    dv_coh = make_dv(f"_ref!$D$1:$D${coh_last}")

    ws.add_data_validation(dv_bt)
    ws.add_data_validation(dv_ts)
    ws.add_data_validation(dv_src)
    ws.add_data_validation(dv_coh)

    dv_bt.sqref = "D2:D1001"
    dv_ts.sqref = "F2:F1001"
    dv_src.sqref = "G2:G1001"
    dv_coh.sqref = "H2:H1001"

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="SalesPilot_Leads_Template.xlsx")


@bulk_bp.post("/upload")
def bulk_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx files supported"}), 400

    # Fetch cohorts for name→id lookup
    cohorts = supabase.table("cohorts").select("id,name,is_future").execute().data
    cohort_map = {c["name"].strip().lower(): c["id"] for c in cohorts}
    future_id = next((c["id"] for c in cohorts if c["is_future"]), None)

    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active

    added = 0
    skipped = 0
    errors = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        if not any(row):
            continue
        name = str(row[0]).strip() if row[0] else ""
        phone = str(row[1]).strip() if row[1] else ""
        if not name or not phone:
            skipped += 1
            errors.append(f"Row {i}: missing name or phone")
            continue

        whatsapp = str(row[2]).strip() if row[2] else phone
        business_type = str(row[3]).strip() if row[3] else ""
        city = str(row[4]).strip() if row[4] else ""
        team_size = str(row[5]).strip() if row[5] else ""
        source = str(row[6]).strip() if row[6] else ""
        cohort_name = str(row[7]).strip() if row[7] else ""

        # Resolve cohort
        cohort_id = cohort_map.get(cohort_name.lower()) if cohort_name else None
        if not cohort_id:
            cohort_id = future_id

        try:
            result = supabase.table("leads").insert({
                "name": name, "phone": phone, "whatsapp": whatsapp,
                "business_type": business_type, "city": city,
                "team_size": team_size, "source": source,
                "score": 40, "score_reason": "Bulk imported lead.",
                "ai_recommendation": "Log a call or WhatsApp to get AI scoring.",
                "archived": False
            }).execute()
            lead_id = result.data[0]["id"]
            if cohort_id:
                supabase.table("cohort_leads").insert({
                    "cohort_id": cohort_id, "lead_id": lead_id,
                    "standing": "Interested", "status": "active"
                }).execute()
            added += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Row {i}: {str(e)[:80]}")

    return jsonify({"added": added, "skipped": skipped, "errors": errors[:10]})
